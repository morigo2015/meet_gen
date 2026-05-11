from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .errors import ExcelReadError
from .formatters import is_empty_value, normalize_header


@dataclass(frozen=True)
class ExcelData:
    headers: list[str]
    rows: list[dict[str, Any]]
    sheet_name: str


def read_excel(path: str | Path, sheet_name: str | None = None) -> ExcelData:
    source = Path(path)
    if not source.exists():
        raise ExcelReadError(f"Excel file not found: {source}")
    if source.suffix.lower() != ".xlsx":
        raise ExcelReadError(f"Unsupported Excel file format: {source.suffix or '<none>'}")

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ExcelReadError(f"Invalid Excel file: {source}") from exc
    except Exception as exc:  # pragma: no cover - openpyxl exposes several parser errors.
        raise ExcelReadError(f"Could not read Excel file: {source}") from exc

    try:
        if sheet_name is None:
            if not workbook.sheetnames:
                raise ExcelReadError("Excel file does not contain worksheets")
            worksheet = workbook[workbook.sheetnames[0]]
        else:
            if sheet_name not in workbook.sheetnames:
                raise ExcelReadError(f'Worksheet not found: "{sheet_name}"')
            worksheet = workbook[sheet_name]

        row_iter = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(row_iter)
        except StopIteration as exc:
            raise ExcelReadError("Excel file is empty") from exc

        indexed_headers = _normalize_headers(raw_headers)
        headers = [header for _, header in indexed_headers]
        if not headers:
            raise ExcelReadError("Excel file does not contain column headers")
        _validate_unique_headers(headers)

        rows: list[dict[str, Any]] = []
        for offset, values in enumerate(row_iter, start=2):
            if _is_empty_row(values):
                continue
            row: dict[str, Any] = {"_row_number": offset}
            for index, header in indexed_headers:
                row[header] = values[index] if index < len(values) else None
            rows.append(row)

        return ExcelData(headers=headers, rows=rows, sheet_name=worksheet.title)
    finally:
        workbook.close()


def _normalize_headers(raw_headers: tuple[Any, ...]) -> list[tuple[int, str]]:
    indexed_headers: list[tuple[int, str]] = []
    for index, value in enumerate(raw_headers):
        header = normalize_header(value)
        if header:
            indexed_headers.append((index, header))
    return indexed_headers


def _validate_unique_headers(headers: list[str]) -> None:
    seen: set[str] = set()
    duplicates: list[str] = []
    for header in headers:
        if header in seen and header not in duplicates:
            duplicates.append(header)
        seen.add(header)
    if duplicates:
        duplicate_list = ", ".join(f'"{item}"' for item in duplicates)
        raise ExcelReadError(f"Duplicate column headers found: {duplicate_list}")


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(is_empty_value(value) for value in values)
